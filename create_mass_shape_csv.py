# export_id_shape_from_mass.py
# דרישות: pip install pandas openpyxl

from pathlib import Path
import re
import datetime as dt
import pandas as pd
from openpyxl import load_workbook

# ========= שנה כאן את הנתיבים =========
XLSX_PATH = Path("/mnt/d/DATA_CMMD/manifest-1734116293719/TOMPEI-CMMD_clinical_data_v01_20250121.xlsx")
SHEET_NAME_HINT = "Imaging Diagnosis Details Sheet"   # הגיליון השני בתמונה
OUT_XLSX = Path("/mnt/d/DATA_CMMD/manifest-1734116293719/mass_id_shape_output11.xlsx")
# =====================================

# --------- עזר: קריאה ונירמול טקסט תאים ---------
def _cell_str(x):
    return str(x).strip() if x is not None else ""

def _norm_pid(s: str) -> str | None:
    """
    חילוץ מזהה בפורמט D?-dddd (למשל D1-0123) מתא כלשהו.
    רגיש לרווחים/NO-BREAK SPACE ומנרמל ל-UPPER.
    """
    s = _cell_str(s).upper().replace("\u00A0", " ").strip()
    m = re.search(r"\bD\d-\d{4}\b", s)
    return m.group(0) if m else None

def _norm_shape(x) -> str | None:
    if x is None or pd.isna(x):
        return None
    s = str(x).strip()
    if not s or s.lower() == "nan":
        return None
    return s

def _norm_leftright(x) -> str | None:
    """
    ממפה ערכים שונים ל-L/R; אם כבר L/R—נשאיר.
    """
    if x is None or pd.isna(x):
        return None
    s = str(x).strip()
    if not s or s.lower() == "nan":
        return None
    t = s.upper().replace(".", "").replace("-", "").replace(" ", "")
    if t in {"L", "R"}:
        return t
    if t in {"LEFT", "LT", "שמאל"}:
        return "L"
    if t in {"RIGHT", "RT", "ימין"}:
        return "R"
    if "/" in t:  # לדוגמה L/R
        parts = [p for p in t.split("/") if p]
        if parts and parts[0] in {"L", "R"}:
            return parts[0]
    return None

# --------- עזר: איתור כותרות (כולל כותרות-על ממוזגות) ---------
def _find_all(ws, text, max_rows=120):
    """
    חיפוש כל התאים ששווים לטקסט (לא רגיש לאותיות גדולות/קטנות).
    מחזיר רשימת קואורדינטות 1-based: [(row, col), ...]
    """
    hits = []
    t = text.strip().lower()
    for r in range(1, min(ws.max_row, max_rows) + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip().lower() == t:
                hits.append((r, c))
    return hits

def _is_mass_above_same_col(ws, row, col, window=8):
    for r in range(max(1, row - window), row):
        v = ws.cell(row=r, column=col).value
        if isinstance(v, str) and v.strip().lower() == "mass":
            return True
    return False

def _is_under_mass(ws, row, col, vspan=8, hspan=12):
    """
    מחזיר True אם התא (row,col) נמצא תחת כותרת-על 'Mass'.
    """
    if _is_mass_above_same_col(ws, row, col, vspan):
        return True
    for m in ws.merged_cells.ranges:
        if m.min_row <= (row - 1) <= m.max_row and m.min_col <= col <= m.max_col:
            top_left = ws.cell(m.min_row, m.min_col).value
            if isinstance(top_left, str) and top_left.strip().lower() == "mass":
                return True
    for r in range(max(1, row - vspan), row):
        for c in range(max(1, col - hspan), col + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip().lower() == "mass":
                return True
    return False

# --------- קריאה מהאקסל ובניית הטבלה ID, SHAPE, LEFTRIGHT ---------
def extract_id_shape_lr(xlsx_path: Path, sheet_name_hint: str) -> pd.DataFrame:
    wb = load_workbook(xlsx_path, data_only=True)

    # בחירת גיליון
    if sheet_name_hint in wb.sheetnames:
        ws = wb[sheet_name_hint]
    elif len(wb.worksheets) >= 2:
        ws = wb.worksheets[1]
    else:
        ws = wb.active

    # איתור עמודת ה-ID (טופ־לבל)
    id_hits = _find_all(ws, "ID")
    if not id_hits:
        id_hits = _find_all(ws, "Patient ID")
    if not id_hits:
        raise ValueError("לא נמצאה כותרת ID או 'Patient ID'.")
    id_r, id_c = id_hits[0]

    # איתור עמודת ה-Shape תחת Mass (כמו קודם)
    shape_candidates = _find_all(ws, "Shape")
    shape_hits = [hc for hc in shape_candidates if _is_under_mass(ws, hc[0], hc[1])]
    if not shape_hits:
        raise ValueError("לא נמצאה עמודת 'Shape' תחת הכותרת 'Mass'.")
    shape_r, shape_c = shape_hits[0]

    # *** חדש: LEFTRIGHT היא כותרת־טופ (לא תחת Mass), בדיוק בשם הזה ***
    lr_hits = _find_all(ws, "LEFTRIGHT")
    if not lr_hits:
        raise ValueError("לא נמצאה כותרת 'LEFTRIGHT'.")
    lr_r, lr_c = lr_hits[0]

    # קוראים את הגיליון ל־DataFrame ללא headers כדי לשלוט באינדקסים
    df = pd.read_excel(xlsx_path, sheet_name=ws.title, header=None, engine="openpyxl")

    # תחילת הדאטה: השורה אחרי כל כותרת שמצאנו
    start = max(id_r, shape_r, lr_r)  # ws הוא 1-based; df הוא 0-based ולכן אין +1
    id_series = df.iloc[start:, id_c - 1]
    shape_series = df.iloc[start:, shape_c - 1]
    lr_series = df.iloc[start:, lr_c - 1]

    rows = []
    for raw_id, raw_shape, raw_lr in zip(id_series, shape_series, lr_series):
        pid = _norm_pid(raw_id)
        shp = _norm_shape(raw_shape)
        lr_val = _norm_leftright(raw_lr) or ""
        if pid and shp:
            rows.append((pid, shp, lr_val))

    return pd.DataFrame(rows, columns=["ID", "SHAPE", "LEFTRIGHT"])

# --------- כתיבה ל-XLSX כולל README ---------
def write_output_with_readme(df: pd.DataFrame, out_xlsx: Path, src_path: Path, sheet_used: str):
    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        # גיליון התוצאה
        df.to_excel(writer, sheet_name="ID_SHAPE", index=False)

        # גיליון README
        meta = {
            "Generated at": [dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            "Source file": [str(src_path)],
            "Sheet used": [sheet_used],
            "Extraction rule": [
                "ID and LEFTRIGHT from top-level headers; SHAPE from under the 'Mass' super-header"
            ],
            "ID pattern": [r"D\d-\d{4}"],
            "Output file": [str(out_xlsx)],
            "Columns": ["ID, SHAPE, LEFTRIGHT"],
            "LEFTRIGHT header location": ["Top-level (same header row as ID)"],
            "Notes": [
                "Rows without a valid ID or empty Shape are skipped. LEFTRIGHT normalized to L/R; empty if missing."
            ],
        }
        pd.DataFrame(meta).to_excel(writer, sheet_name="README", index=False)

def main():
    df = extract_id_shape_lr(XLSX_PATH, SHEET_NAME_HINT)
    if df.empty:
        print("לא נמצאו רשומות ID,SHAPE.")
    else:
        print(f"נמצאו {len(df)} רשומות עם ID, SHAPE ו-LEFTRIGHT.")
    # לשם התיעוד בתוך הקובץ, נחלץ את שם הגיליון שבפועל שימש:
    wb = load_workbook(XLSX_PATH, data_only=True)
    sheet_used = SHEET_NAME_HINT if SHEET_NAME_HINT in wb.sheetnames else (
        wb.worksheets[1].title if len(wb.worksheets) >= 2 else wb.active.title
    )
    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    write_output_with_readme(df, OUT_XLSX, XLSX_PATH, sheet_used)
    print(f"הקובץ נשמר: {OUT_XLSX}")

if __name__ == "__main__":
    main()
